from tkinter import ttk
import tkinter as tk
import openpyxl
from tkinter import messagebox
from datetime import date
import pandas as pd
data_atual = date.today()


def atualizar_combobox(event):

    produto_selecionado = produto_combobox.get()   
    codigos = codigos_por_veiculo.get(produto_selecionado, [''])
    tamanhos = tamanho_por_veiculo.get(produto_selecionado,[''])

    tamanho_combobox['values'] = tamanhos
    modelo_combobox['values'] = codigos
    modelo_combobox.current(0) 
    tamanho_combobox.current(0)

def verificar_input(dados):
    if isinstance(dados, int) or isinstance(dados,float): 
        return True
    if dados.strip() == '': 
        return False
    return True

def cadastrar_produto():

    try:
        dia=int(dia_combobox.get())
        if dia<=0 or dia>31:
            raise ValueError
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, Insira uma Data válida.")
        return

    try:
        mes=int(mes_combobox.get())
        if mes<=0 or mes>12:
            raise ValueError
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, Insira uma Data válida.")
        return
    mes_nome, quarter = atributo_mes[str(mes)]

    try:
        ano=int(anos_combobox.get())
        if ano<=200 or ano>2050:
            raise ValueError
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, Insira uma Data válida.")
        return

    data=(f"{dia}/{mes}/{ano}")

    try:
        codigo = int(codigo_entry.get())
        if codigo <= 0:
            raise ValueError("Por favor, digite um código válido.")
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, insira um valor válido para Código.")
        return

    try:
        quantidade = int(quantidade_spinbox.get())
        if quantidade <= 0:
            raise ValueError("Por favor, digite uma quantidade válida.")
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, insira um valor válido para quantidade.")
        return

    try:
        preco = float(preco_entry.get())
        if preco <= 0:
            raise ValueError("Por favor, digite um preço válido.")
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, insira um valor válido para o Preço.")
        return

    try:
        num = int(num_entry.get())
        if num <= 0:
            raise ValueError("Por favor, digite um número de produto válido.")
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, insira um valor válido para o ID do produto.")
        return

    try:
        cep = int(cep_entry.get())
        if cep <= 0:
            raise ValueError("Por favor, digite um CEP válido.")
    except ValueError:
        messagebox.showerror("Input Error", "Por favor, insira um valor válido para o CEP.")
        return

    try:
        status = status_combobox.get()
        if status=='':
            raise ValueError
    except:
        messagebox.showerror('Input Error','Por favor, insira um valor válido para Status')
        return

    try:
        produto = produto_combobox.get()
        if produto=='':
            raise ValueError
    except:
        messagebox.showerror('Input Error','Por favor, insira um valor válido para o Produto')
        return
    
    try:
        modelo = modelo_combobox.get()
        if modelo == '':
            raise ValueError
    except:
        messagebox.showerror('Input Error','Por favor, insira um valor válido para o Modelo do veículo')
        return
    
    try:
        estado = estado_combobox.get()
        if estado=='':
            raise ValueError
    except:
        messagebox.showerror('Input Error','Por favor, insira um valor válido para o seu Estado')
        return
    
    try:
        tamanho = tamanho_combobox.get()
        if tamanho=='':
            raise ValueError
    except:
        messagebox.showerror('Input Error','Por favor, insira um valor válido para o Tamanho do veículo')
        return

    # dados=[data,codigo,quantidade,preco,num,vendas,status,produto,modelo,estado,cep,tamanho]
    # for i in dados:
    #      if not (verificar_input(i)):
    #         messagebox.showerror("Input Error", "Todos os campos devem ser preenchidos.")
    #         return

    df = pd.read_excel('estoque.xlsx') 
    df_produtos = df['Produto']   
    produto_filtrado = df[df_produtos == produto]
    quantidadeProduto = produto_filtrado['Quantidade'].values[0]
    quantidade_estoque = int(quantidadeProduto)
    quantidade = int(quantidade)

    if quantidade > quantidade_estoque:
        messagebox.showerror('Produto no Estoque',f'Estoque do {produto} em falta!')
        return None
    else:
        df.loc[df['Produto'] == produto, 'Quantidade'] -= quantidade    
        df.to_excel('estoque.xlsx', index=False)

         
    planilha='sales.xlsx'
    workbook = openpyxl.load_workbook(planilha)
    folha = workbook.active 
    id_prod=folha.max_row - 1
    faturamento = quantidade*preco
    info_produto=[id_prod,codigo,quantidade,preco,num,data,status,quarter,mes_nome,ano,produto,modelo,estado,cep,'EUA',tamanho, faturamento, faturamento]
    folha.append(info_produto)
    workbook.save(planilha)
    messagebox.showinfo('Cadastro de venda','Venda cadastrada com sucesso!')
    treeview.insert('',tk.END,values=info_produto)

    dia_combobox.set(data_atual.day)
    mes_combobox.set(data_atual.month)
    anos_combobox.set(2024)
    codigo_entry.delete('0','end')
    quantidade_spinbox.delete('0','end')
    preco_entry.delete('0','end')
    num_entry.delete('0','end')
    cep_entry.delete('0','end')
    status_combobox.set('')
    produto_combobox.set('')
    modelo_combobox.set('')
    estado_combobox.set('')
    tamanho_combobox.set('')
            

def carregar_planilha():
    planilha='sales.xlsx'
    workbook = openpyxl.load_workbook(planilha)
    folha = workbook.active

    lista_valores=list(folha.values)

    treeview['columns'] = lista_valores[0]
    for col in lista_valores[0]:
        treeview.heading(col, text=col)
        treeview.column(col, width=100)
    
    for row in lista_valores[1:]:
        if all(cell is not None for cell in row):
            treeview.insert('', tk.END, values=row)

def finalizar():
        root.destroy()





root = tk.Tk()
# root.tk.call('source', "azure.tcl")
# root.tk.call('set_theme','light')
root.title('Painel de vendas')
root.geometry('1920x1080')
frame = ttk.Frame(root)
frame.pack()


# QUADRANTE PARA AS INFORMACOES DO PEDIDO
input_frame = ttk.LabelFrame(frame, text='Cadastro de venda',height=1080)
input_frame.grid(row=0, column=0,padx=20,pady=100)
input_frame.pack_propagate(False)

sy=ttk.Scrollbar(input_frame)



# CAIXA DE TEXTO PARA CODIGO DO PEDIDO
codigo_entry = ttk.Entry(input_frame)
ttk.Label(input_frame, text='Codigo do pedido:').grid(row=0,column=0)
# codigo_entry.insert(0,'insira o codigo da venda')                                 # INSERE UM ELEMENTO NA CAIXA DE TEXTO 
# codigo_entry.bind('<FocusIn>', lambda x: codigo_entry.delete('0', 'end'))         # DELETA O TEXTO AO CLICAR NA CAIXA DE TEXTO
codigo_entry.grid(row=0, column=1, sticky='ew',padx=5,pady=10)


# SPINBOX PARA QUANTIDADE
quantidade_spinbox = ttk.Spinbox(input_frame, from_=1, to=999)
ttk.Label(input_frame, text='Quantidade:').grid(row=1,column=0)
quantidade_spinbox.grid(row=1, column=1,sticky='ew',padx=5,pady=10)


# ENTRY CAIXA DE TEXTO PARA PREÇO
preco_entry=ttk.Entry(input_frame)
ttk.Label(input_frame, text='Preço:').grid(row=2,column=0) 
preco_entry.grid(row=2, column=1, sticky='ew',padx=5,pady=10)


# ENTRY CAIXA DE TEXTO PARA NUMERO DO PRODUTO (ID DO PRODUTO)
num_entry=ttk.Entry(input_frame)
ttk.Label(input_frame, text='Número do produto:').grid(row=3,column=0)  
num_entry.grid(row=3, column=1, sticky='ew',padx=5,pady=10)


atributo_mes = {
    '1': ['Jan','1'],
    '2': ['Feb','1'],
    '3': ['Mar','1'],
    '4': ['Apr','2'],
    '5': ['May','2'],
    '6': ['Jun','2'],
    '7': ['Jul','3'],
    '8': ['Ago','3'],
    '9': ['Sep','3'],
    '10': ['Oct','4'],
    '11': ['Nov','4'],
    '12': ['Dec','4']
}

# DROPBOX PARA DATA
dias = list(range(1,32))
dia_combobox = ttk.Combobox(input_frame, values=dias)
ttk.Label(input_frame, text='Data:').grid(row=5,column=0)
dia_combobox.current(data_atual.day-1)
dia_combobox.grid(row=5, column=1,sticky='ew',padx=5,pady=10)

meses = list(range(1,13))
mes_combobox = ttk.Combobox(input_frame, values=meses)
mes_combobox.current(data_atual.month-1)
mes_combobox.grid(row=5, column=2,sticky='ew',padx=5,pady=10)

anos = list(range(2000,2050))
anos_combobox = ttk.Combobox(input_frame, values=anos)
anos_combobox.current(24)
anos_combobox.grid(row=5, column=3,sticky='ew',padx=5,pady=10)



# DROP BOX PARA O STATUS DO PEDIDO
statuses=['','Shipped', 'In Process', 'On Hold', 'Resolved', 'Cancelled']
status_combobox = ttk.Combobox(input_frame, values=statuses)
ttk.Label(input_frame, text='Status:').grid(row=6,column=0)
status_combobox.current(0)
status_combobox.grid(row=6, column=1,sticky='ew',padx=5,pady=10)



# DICIONARIO PARA OS MODELOS DE CADA VEICULO
codigos_por_veiculo = {
    'Motorcycles': ['S10_1678', 'S10_2016', 'S10_4698', 'S12_2823', 'S18_2625', 'S18_3782'],
    'Classic Cars': ['S10_1949', 'S10_4757', 'S10_4962', 'S12_1099', 'S12_1108', 'S12_3148', 'S12_3380', 'S12_3891', 'S12_3990', 'S12_4675', 'S18_1129', 'S18_1589', 'S18_1889', 'S18_1984', 'S18_2238', 'S18_2870', 'S18_3232', 'S18_3278', 'S18_3482', 'S18_4027', 'S18_4721', 'S18_4933'],
    'Trucks and Buses': ['S12_1666', 'S12_4473', 'S18_1097', 'S18_2319', 'S18_2432', 'S18_4600'],
    'Vintage Cars': ['S18_1342', 'S18_1342', 'S18_1367', 'S18_1749', 'S18_2248', 'S18_2325', 'S18_2795', 'S18_2949', 'S18_2957', 'S18_3136', 'S18_3140', 'S18_3320', 'S18_3856', 'S18_4409', 'S18_4522', 'S18_4668'],
    'Planes': ['S18_1662', 'S18_2581'],
    'Ships': ['S18_3029'],
    'Trains': ['S18_3259']
}


# DICIONARIO PARA TAMANHO DOS VEICULOS
tamanho_por_veiculo = {
    'Motorcycles': ['Small','Medium','Large'],
    'Classic Cars': ['Small','Medium','Large'],
    'Trucks and Buses': ['Small','Medium','Large'],
    'Vintage Cars': ['Small','Medium','Large'],
    'Planes': ['Medium','Large'],
    'Ships': ['Large'],
    'Trains': ['Large']
}


# COMBOBOX PARA PRODUTO
produtos=['','Motorcycles', 'Classic Cars', 'Trucks and Buses', 'Vintage Cars', 'Planes', 'Ships', 'Trains']
produto_combobox = ttk.Combobox(input_frame, values=produtos)
ttk.Label(input_frame, text='Veiculo:').grid(row=7,column=0)
produto_combobox.current(0)
produto_combobox.grid(row=7, column=1,sticky='ew',padx=5,pady=10)


# COMBOBOX PARA O CODIGO DO PRODUTO (MODELO DO VEICULO)
modelo_combobox = ttk.Combobox(input_frame, values=[''])
ttk.Label(input_frame, text='Modelo:').grid(row=8,column=0)
modelo_combobox.current(0)
modelo_combobox.grid(row=8, column=1,sticky='ew',padx=5,pady=10)
produto_combobox.bind('<<ComboboxSelected>>',atualizar_combobox)


# TAMANHO DO VEÍCULO
tamanho_combobox = ttk.Combobox(input_frame, values=[''])
ttk.Label(input_frame, text='Tamanho:').grid(row=9,column=0)
tamanho_combobox.current(0)
tamanho_combobox.grid(row=9, column=1,sticky='ew',padx=5,pady=10)


# COMBOBOX PARA ESTADO
estados=['','AL','AK','AZ','AR','CA','NC','SC','CO','CT','ND','SD','DE','DC','FL','GA','ID','IL','IN','IA','KS','KY','LA','ME','MD','MA','MI','MN','MS','MO','MT','NE','NV','NH','NJ','NY']
estado_combobox = ttk.Combobox(input_frame, values=estados)
ttk.Label(input_frame, text='Estado:').grid(row=11,column=0)
estado_combobox.current(0)
estado_combobox.grid(row=11, column=1,sticky='ew',padx=5,pady=10)


# ENTRY CAIXA DE TEXTO PARA CODIGO POSTAL
cep_entry=ttk.Entry(input_frame)
ttk.Label(input_frame, text='Código Postal:').grid(row=10,column=0) 
cep_entry.grid(row=10, column=1, sticky='ew',padx=5,pady=10)



# BOTAO PARA INSERIR DADOS À TABELA
botao = ttk.Button(input_frame, text='Cadastrar',command=cadastrar_produto)
botao.grid(row=12, column=0, sticky='nsew',padx=5,pady=15)

# BOTAO PARA FECHAR JANELA
sair = ttk.Button(input_frame, text='Sair',command=finalizar)
sair.grid(row=12, column=1, sticky='nsew',padx=5,pady=15)




# QUADRANTE PARA SEÇAO PARA AS VENDAS
frame_vendas = ttk.Frame(frame)
frame_vendas.grid(row=0,column=1,pady=10)

scrolly = ttk.Scrollbar(frame_vendas,orient='vertical')
scrolly.pack(side='right',fill='y')

scrollx = ttk.Scrollbar(frame_vendas,orient='horizontal')
scrollx.pack(side='bottom',fill='x')

# SEÇÃO DE VENDAS
colunas = ('ID', 'Pedido', 'Quantidade', 'Preço', 'Numero do Produto', 'Vendas','Data', 'Status', 'Quarter', 'Mês', 'Ano', 'Produto', 'Código do Produto', 'Estado', 'Código Postal', 'Pais', 'Tamanho', 'Faturamento')
treeview=ttk.Treeview(frame_vendas, show='headings', yscrollcommand=scrolly.set, xscrollcommand=scrollx.set, columns=colunas, height=28)
treeview.pack(side='left')
scrolly.config(command=treeview.yview)
scrollx.config(command=treeview.xview)
carregar_planilha()


root.mainloop()