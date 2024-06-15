from PySide6.QtCore import Qt
from PySide6.QtWidgets import *
from login import Ui_Login
from windowMain import Ui_MainWindow
import sys
import subprocess
import time as t
from database import *
import openpyxl

class Login(QWidget,Ui_Login):
    def __init__(self):
        super(Login,self).__init__()
        self.setupUi(self)
        self.setWindowTitle('Login No Sistema')
        self.tentativas = 3

        self.btn_entrar.clicked.connect(self.open_system)

    def open_system (self):

        self.users = Database()
        self.users.connect()

        autenticado = self.users.check_user(self.txt_login.text().upper(), self.txt_password.text().upper())

        if autenticado.lower() == 'adminstrador' or autenticado.lower() == 'usuário':
            self.w = MainWindow(autenticado)
            self.w.show()
            self.close()

        else:
            if self.tentativas > 0:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle('Usuário Inválido')
                msg.setText(f"Usuário Não registrado\n \n Tentativas: {self.tentativas}")
                self.tentativas -=1
                msg.exec()
            
            elif self.tentativas == 0:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle('Sistema Bloqueado')
                msg.setText("O sistema foi Bloqueado, numeros de tentativas excedidas")
                msg.exec()
                t.sleep(3)
                sys.exit()

        self.users.close_connection()

class MainWindow(QMainWindow,Ui_MainWindow):
    def __init__(self,user):
        super(MainWindow,self).__init__()
        self.setupUi(self)
        self.setWindowTitle('Sistema de Vendas')
        

        if user.lower() == 'Usuário':
            self.btn_pg_venda.setVisible(False)
        
        
        self.btn_home.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_home))
        self.btn_pg_venda.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_cadastro))
        self.btn_tables.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_table))
        self.btn_dashboards.clicked.connect(self.subprocess_dash)
        self.btn_cadastro.clicked.connect(self.subscribe_user)
        self.btn_home.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_home))
        self.btn_tables.clicked.connect(self.show_table)
        self.btn_tables.clicked.connect(self.table_estoque)
        self.insert_saida.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_inserir_saida))
        self.btn_saida.clicked.connect(self.cadastrar_venda)

    def subprocess_dash(self):
        subprocess.Popen(['streamlit', 'run', r'Estoque\dashs.py'])

    
    def show_table(self):
        df = pd.read_excel('sales.xlsx')
        colunas_desejadas = ['Pedido', 'Quantidade', 'Produto', 'Preco', 'Faturamento', 'Data']
        df = df[colunas_desejadas]
        values = df.values.tolist() 

        self.tb_saida.clearContents()
        self.tb_saida.setRowCount(len(values))
        self.tb_saida.setColumnCount(len(df.columns))

        self.tb_saida.setHorizontalHeaderLabels(df.columns)

        for row_index, row_data in enumerate(values):
            for column_index, cell_data in enumerate(row_data):
                self.tb_saida.setItem(row_index, column_index, QTableWidgetItem(str(cell_data)))

    def table_estoque(self):
        df = pd.read_excel('estoque.xlsx')
        values = df.values.tolist()

        self.td_estoque.clearContents()
        self.td_estoque.setRowCount(len(values))
        self.td_estoque.setColumnCount(len(df.columns))

        self.td_estoque.setHorizontalHeaderLabels(df.columns)

        for row_index, row_data in enumerate(values):
            for column_index, cell_data in enumerate(row_data):
                self.td_estoque.setItem(row_index, column_index, QTableWidgetItem(str(cell_data)))

        for linha in range(1,6):
            self.td_estoque.resizeColumnsToContents()

    def cadastrar_venda(self):
        produto = self.txt_produto.text().capitalize()
        quantidade = self.txt_quantidade.text()
        data = self.cb_data.text()
        preco = self.txt_preco.text()
        vazio = ''
        df = pd.read_excel('estoque.xlsx') 
        df_produtos = df['Produto']

        if produto == vazio or quantidade == vazio or preco == vazio:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle('Valores Inseridos')
            msg.setText('Por Favor, Informe todos os Valores!')
            msg.exec()
            return None
        else:

            if produto not in df_produtos.values:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle('Produto Inválido')
                msg.setText('Produto Não Encontrado\nna Planilha de Estoque!')
                msg.exec()
                return None
            
            produto_filtrado = df[df_produtos == produto]
            quantidadeProduto = produto_filtrado['Quantidade'].values[0]
            quantidade_estoque = int(quantidadeProduto)
            quantidade = int(quantidade)

            if quantidade > quantidade_estoque:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle('Falha na Quantidade')
                msg.setText(f'Quantidade em Estoque do {produto} em Falta! ')
                msg.exec()
                return None
            else:
                df.loc[df['Produto'] == produto, 'Quantidade'] -= quantidade
                
                df.to_excel('estoque.xlsx', index=False)

                arquivo = 'sales.xlsx'
                panilha = openpyxl.load_workbook(arquivo)
                folha = panilha.active

                
                proxima_linha = folha.max_row + 1

                
                folha.cell(column=11, row=proxima_linha, value=produto)  
                folha.cell(column=3, row=proxima_linha, value=quantidade) 
                folha.cell(column=6, row=proxima_linha, value=data)  
                folha.cell(column=4, row=proxima_linha, value=preco)  

                panilha.save(arquivo)

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle('Cadastro Realizado')
                msg.setText('Venda Cadastrada Com Sucesso')
                msg.exec()


    def cadastrar_produto(self):
        pass

    def subscribe_user(self):

        if self.txt_senha.text() != self.txt_senha2.text():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Senhas Diferentes")
            msg.setText("A senha não é igual")
            msg.exec()
            return None
        
        else:
            name = self.txt_nome.text()
            user = self.txt_usuario.text()
            password = self.txt_senha.text()
            access = self.cb_perfil.currentText()
            
            db = Database()
            db.connect()
            try:
                db.insert_user(name,user,password,access)
                msgs = QMessageBox()
                msgs.setIcon(QMessageBox.Information)
                msgs.setWindowTitle('Cadastrado Efetivado')
                msgs.setText("Usuário Cadastrado")
                msgs.exec()
            except sqlite3.IntegrityError:
                msg2 = QMessageBox()
                msg2.setIcon(QMessageBox.Information)
                msg2.setWindowTitle("Usuário Cadastrado")
                msg2.setText(F"O Usuário {name} já Está Cadastrado!")
                msg2.exec()
                return None
            finally:
                db.close_connection()
            
        
if __name__=='__main__':
    app = QApplication(sys.argv)
    window = Login()
    window.show()
    app.exec()
    